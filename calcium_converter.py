import argparse
from statistics import mean, stdev
from textwrap import indent
from typing import Dict, List, Tuple, NewType, Union
import os
import sys
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class PeakValues:
    '''
    Possible input values for peak calculations.
    '''
    PEAK_HIGHEST_VALUE = 1
    PEAK_AVERAGE_OF_THREE = 2
    VALID_PEAK_SELECTIONS = [PEAK_HIGHEST_VALUE, PEAK_AVERAGE_OF_THREE]


class SheetParseValues:
    # Row number (1-indexed) where the header is located
    HEADER_ROW = 1
    # Row number (1-indexed) where data starts in the calcium imaging spreadsheet.
    START_DATA_ROW = 2
    # Row to start printing calculations to
    CALCULATIONS_START_ROW = 3
    # Column to the right of the data to start printing calculations to
    CALCULATIONS_START_COL = 4
    # Column that treatment labels are written
    TREATMENT_LABEL_COLUMN = 2


class TimeValuePair:    
    def __init__(self, time, value):
        self.time = time
        self.value = value

    def __str__(self) -> str:
        return f'({self.time}, {self.value})'


class TreatmentData:
    def __init__(self):
        self.treatment_name: str = ""
        self.prior_wash_data: List[TimeValuePair] = list()
        self.treatment_data: List[TimeValuePair] = list()
        self.anterior_wash_data: List[TimeValuePair] = list()


class StandardBathData:
    def __init__(self):
        self.std_bath_data: List[TimeValuePair] = list()


class LabelIndexPair:
    def __init__(self, label: str, index: int):
        self.label = label
        self.index = index

    def __str__(self):
        return f'({self.label}, {self.index})'


CalciumDataRegionToTreatmentsDict = NewType('CalciumDataRegionToTreatmentsDict', Dict[str, List[Union[TreatmentData, StandardBathData]]])
TreatmentLabels = NewType('TreatmentLabels', List[LabelIndexPair])


def is_valid_treatment(potential_treatment: str) -> bool:
    '''
    @return True if the potential treatment label is valid, false otherwise.
    '''
    return 'Standard Bath' != potential_treatment and \
           'STD' != potential_treatment


def is_valid_ratio_data_header(potential_header: str) -> bool:
    '''
    @return True if the potential header is a valid ratio data label, false
    otherwise.
    '''
    return 'Ratio' in potential_header


def infile_valid(infile: str) -> str:
    '''
    Ensures the input filename is a valid file that can be read.
    
    Trys to open with openpyxl.

    @param infile: Input filename to check.
    '''
    try:
        _ = openpyxl.load_workbook(infile)
        return True
    except FileNotFoundError as e:
        print(e)
        return False
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(e)
        return False


def parse_arg() -> argparse.Namespace:
    '''
    Parses cli arguments, returning argparse Namespace with arguments parsed.

    Will exit upon argument parse failure.
    '''
    parser = argparse.ArgumentParser(description='')
    parser.add_argument(
        'file',
        type=str,
        help='Calcium file to convert. Must be an xlsx file with a name of the form'
             'XXXX_XX_XX_Y.xlsx, where XXXX_XX_XX is a date and Y is a run label.')
    parser.add_argument(
        '-base',
        type=int,
        default=10,
        help='Number of cycles before drug application')
    parser.add_argument(
        '-peak',
        type=int,
        default=1,
        help='Enter (1) if you want to calculate the PEAK using the highest value,'
             '(2) if using the average of three')
    parser.add_argument(
        '--post-std-time-to-search',
        type=int,
        default=5*60,
        help='Number of seconds after treatment applied to search for a peak')

    args = parser.parse_args()

    if args.peak not in PeakValues.VALID_PEAK_SELECTIONS:
        parser.print_help()
        parser.exit('\ninvalid peak selection')

    if not infile_valid(args.file):
        parser.exit('invalid input filename')

    return args


def parse_data_name(file: str) -> dict:
    '''
    Parses a calcium image filename. The filename contains a date and run
    letter, which will be extracted and returned.

    @param file: Filename of the calcium image file
    '''
    date_length             = len('XXXX_XX_XX')
    date_run_divider_length = len('_')
    run_length              = len('Y')
    xlsx_length             = len('.xlsx')

    run_index = date_length + date_run_divider_length

    # Remove path to file if it exists
    file = os.path.split(file)[1]

    name_args = dict()

    if len(file) < date_length + date_run_divider_length + run_length + xlsx_length:
        print('Invalid file name, must have name XXXX_XX_XX_Y.xlsx, where XXXX_XX_XX '
              'is a date and Y is a run (A, B, C, etc.)')
        sys.exit(1)

    name_args['date'] = file[:date_length]
    name_args['run'] = file[run_index]

    return name_args


def generate_outfilename(infile: str) -> str:
    '''
    Generates outfilename based on infile's name.

    Generated file will have format {infilename}_analysis_new{infileext}

    @param infile: Input filename.
    '''
    infilename, infileext = os.path.splitext(infile)
    filelocation, infilename = os.path.split(infilename)
    return os.path.join(filelocation, infilename) + '_analysis' + infileext


def convert_calcium_file(
    infile: str,
    outfile: str,
    run_label: str,
    base_elements: int,
    peak_type: int,
    post_std_time_to_search: int) -> None:
    '''
    Performs conversions on calcium imaging file, storing resulting conversion
    in output file.
    
    @param infile: Input excel file. @param outfile: Output filename. @param
    run_label The run of the day (A, B, C, D, etc.). false otherwise. @param
    base_elements Number of cycles before the application of the drug
    application. @param peak_type 1 if you want to calculate peak using the
    highest value, 2 if you want to calculate peak using an average of 3 values.
    @param post_std_time_to_search: Number of seconds after treatment to search
    for a peak
    '''
    in_wb = openpyxl.load_workbook(infile)
    in_sheet = in_wb.worksheets[0]
    in_sheetname = in_wb.sheetnames[0]

    out_wb = openpyxl.Workbook()

    out_sheetname = f'{in_sheetname} analysis'

    out_wb.create_sheet(out_sheetname)

    out_sheet = out_wb.worksheets[out_wb.sheetnames.index(out_sheetname)]

    # Search input sheet for data
    calcium_data, treatment_labels = get_raw_data(in_sheet, run_label)

    # Number of cell regions in the calcium data
    num_regions = len(calcium_data)

    # Write trial treatment_labels to the sheet
    write_treatment_labels(out_sheet, treatment_labels)

    # Write raw data to output sheet
    write_data_to_outsheet(out_sheet, calcium_data)

    for region_idx, (region_name, region_data) in enumerate(calcium_data.items()):
        treatment_idx = 0
        for treatment_data in region_data:
            if type(treatment_data) is TreatmentData:
                # Write calculations header for the given treatment idx
                write_calculations_header(out_sheet, num_regions, treatment_idx, treatment_data.treatment_name)

                # Compute trial run data for each run
                peak, peak_time = calculate_peak(treatment_data, peak_type, post_std_time_to_search)
                base, std = calculate_base(treatment_data, base_elements)
                area = calculate_area(treatment_data, base, std, peak_time)
                delta = calculate_delta(peak, base)

                write_calculations_to_outsheet(
                    out_sheet, peak,
                    base,
                    std,
                    area,
                    delta,
                    num_regions,
                    region_idx,
                    region_name,
                    treatment_idx)

                treatment_idx += 1

    out_wb.save(outfile)

    print(f"SUCCESS! Your file was written to {outfile}")


def calculate_peak(treatment_data: TreatmentData, peak_type: int, post_std_time_to_search) -> Tuple[float, float]:
    '''
    Returns the peak of the current run.
    
    @param treatment_data: List of treatment data from the current run in some
    particular region. Looks 5 minutes past the treatment data into the anterior
    STD data. @param peak_type: 1 if only the top value should be used to
    compute peak, 2 if the mean of the top 3 values should be used as the peak.
    '''
    end_treatment_value_time = treatment_data.treatment_data[-1].time

    i = 0
    for value in treatment_data.anterior_wash_data:
        i += 1
        if value.time > end_treatment_value_time + post_std_time_to_search:
            break

    data_to_search = treatment_data.treatment_data + treatment_data.anterior_wash_data[0:min(len(treatment_data.anterior_wash_data), i)]

    data_to_search = sorted(data_to_search, reverse=True, key=lambda x: x.value)
    return (data_to_search[0].value, data_to_search[0].time) if peak_type == PeakValues.PEAK_HIGHEST_VALUE else (mean([x.value for x in data_to_search[0:2]]), max(data_to_search[0:2]))


def calculate_base(treatment_data: TreatmentData, base: int) -> float:
    '''
    Returns the base of the current run.

    @param base: The number of values to use to compute the base.
    '''
    base_time_value_pairs = treatment_data.prior_wash_data[-base:]
    base_values = [x.value for x in base_time_value_pairs]
    return mean(base_values), stdev(base_values)


def calculate_area(treatment_data: TreatmentData, base: float, base_std: float, base_peak_time: float) -> float:
    '''
    Computes the positive integral of the current trial above the base.

    @param treatment_data: List of data from current treatment for some
    particular region. @param base: Base to compute auc
    '''
    DEFAULT_DT = 6

    auc = 0.0

    values_to_consider = treatment_data.treatment_data + treatment_data.anterior_wash_data

    consec_values_at_base = 0
    NUM_CONSECUTIVE_VALUES_BEFORE_DONE = 15

    previous_time = None
    for value in values_to_consider:
        dt = DEFAULT_DT if previous_time is None else value.time - previous_time 
        previous_time = value.time
        if value.value > base + base_std:
            auc += (value.value - base) * dt
        elif value.time > base_peak_time:
            consec_values_at_base += 1

            if consec_values_at_base >= NUM_CONSECUTIVE_VALUES_BEFORE_DONE:
                break

    return auc


def calculate_delta(peak: float, base: float) -> float:
    '''
    Computes difference between the current trial's peak and the base.

    @param peak: Peak value computed by calculate_peak. @param base: Base value
    computed by calculate_base.
    '''
    return peak - base


def ratio_to_calcium_concentration(ratio: float) -> float:
    '''
    Converts ratio data to calcium concentration

    @param ratio_data: ratio value to convert to concentration data.
    '''
    if ratio is None:
        return None
    
    return 146 * (25813.79 / 1674.68) * ((ratio - 0.132) / (6.274 - ratio))


def find_number(label: str) -> int:
    '''
    Finds and returns first integer in the label.

    If the label is "abc123def456", the number returned is 123.
    '''
    int_str = ''
    int_found = False
    for c in label:
        if int_found and not c.isdigit():
            break
        
        if c.isdigit():
            int_found = True
            int_str += c

    if len(int_str) == 0:
        return -1

    return int(int_str)


def generate_region_label(old_label: str, region_idx: int, run_label) -> str:
    '''
    Returns a new label that incorporates the old label's trial number and the
    cell label. Uses region_idx if no region number found in old label.

    @param old_label "Old" name used to label the current trial. @param
    region_idx The current 0-based index that indicates which label is being
    created. @param run_label The run of the day (A, B, C, D, etc.).
    '''
    number = find_number(old_label)
    if number == -1:
        return f'{run_label}{region_idx}'
    return f'{run_label}{number}'


def search_for_treatment_labels(insheet: Worksheet) -> TreatmentLabels:
    '''
    Finds and returns a list of (label, label index) pairs that define each
    trial.

    A label column contains a column labeled "Labels" at the top of the column.
    Within the column, all values are empty except for the labels.

    @param insheet: Worksheet to find labels in
    '''
    treatment_labels: TreatmentLabels = list()

    # find column labeled "Labels"
    treatment_label_col = None
    for col in insheet.columns:
        if col[0].value is not None and 'Labels' in col[0].value:
            if treatment_label_col is not None:
                print('Found multiple columns with label "Labels"')

            treatment_label_col = col
    
    if treatment_label_col is None:
        print('failed to find column with "Labels" in first row')
        sys.exit(1)

    for item in treatment_label_col:
        # Skip first row since that is the header, skip values that have '/60'
        # indicating the value isn't a treatment but a time value
        if item.row != 1 and item.value is not None and '/60' not in item.value:
            treatment_labels.append(LabelIndexPair(item.value, item.row))

    return treatment_labels
    
    
def get_raw_data(insheet: Worksheet, run_label: str) \
    -> Tuple[CalciumDataRegionToTreatmentsDict, TreatmentLabels]:
    '''
    Finds time and raw calcium ratio data from the input worksheet.

    @param insheet: Worksheet to search for data in. @param run_label: The run
    of the day (A, B, C, D, etc.)

    @return Tuple with the following elements:
        - dictionary. Maps region name -> [list of TreatmentLabels objects].
        - List of (treatment label, treatment label index) pairs.
    '''
    start_of_data_row, time_column_idx, start_region_idx, end_region_idx = search_for_data_bounds(insheet)

    treatment_labels = search_for_treatment_labels(insheet)

    ratio_data: CalciumDataRegionToTreatmentsDict = dict()

    # Iterate through all regions in the sheet
    for region in range(start_region_idx, end_region_idx + 1):
        old_sheet_label = insheet.cell(1, region).value
        region_label = generate_region_label(old_sheet_label, region - start_region_idx, run_label)

        ratio_data[region_label] = list()

        # Incorporate any data not associated with a label by adding "STD" label
        if treatment_labels[0].index > start_of_data_row:
            treatment_labels.insert(0, ('STD', start_of_data_row))

        # For each region, generate a dictionary mapping (label -> list of
        # ratios)
        for j, label_index_pair in enumerate(treatment_labels):
            treatment_end_idx = treatment_labels[j + 1].index if j + 1 < len(treatment_labels) else insheet.max_row + 1

            def append_values(list: List[float], start: int, end: int) -> None:
                for k in range(end - start):
                    ratio_value = ratio_to_calcium_concentration(insheet.cell(k + start, region).value)
                    time_value = insheet.cell(k + start, time_column_idx).value

                    if (type(ratio_value) is int or type(ratio_value) is float) and \
                        (type(time_value) is int or type(time_value) is float):
                        list.append(TimeValuePair(time_value, ratio_value))
                    else:
                        print(f'Invalid time/ratio values: time {time_value} ratio {ratio_value}, row {k + start}')

                        list.append(TimeValuePair(0, 0))

            if is_valid_treatment(label_index_pair.label):
                treatment_data = TreatmentData()
                treatment_data.treatment_name = label_index_pair.label
                
                # Find and add all treatment data
                append_values(treatment_data.treatment_data, label_index_pair.index, treatment_end_idx)

                assert j > 0, f'STD must be first treatment, non STD is treatment {j}'

                # Use last prior STD data as TreatmentData.prior_wash_data
                prior_std_data = None
                for data in ratio_data[region_label]:
                    if type(data) is StandardBathData:
                        prior_std_data = data.std_bath_data

                treatment_data.prior_wash_data = prior_std_data

                # Add anterior STD treatment data if the next treatment is STD
                if not is_valid_treatment(treatment_labels[j + 1].label):
                    # Use 2 treatment labels ahead to find end of anterior std
                    # treatment
                    anterior_std_treatment_end = treatment_labels[j + 2].index if j + 2 < len(treatment_labels) else insheet.max_row + 1

                    # Find and add all anterior STD data
                    append_values(treatment_data.anterior_wash_data, treatment_end_idx + 1, anterior_std_treatment_end)

                ratio_data[region_label].append(treatment_data)
            else:
                bath_data = StandardBathData()

                append_values(bath_data.std_bath_data, label_index_pair.index, treatment_end_idx)

                ratio_data[region_label].append(bath_data)

    return (ratio_data, treatment_labels)


def write_calculations_header(outsheet, num_regions, treatment_idx, treatment_label) -> None:
    calculation_labels = [treatment_label, 'peak', 'base', 'std', 'area', 'delta']

    for i, label in enumerate(calculation_labels):
        outsheet.cell(
            SheetParseValues.CALCULATIONS_START_ROW + treatment_idx * (len(calculation_labels) + 1) + i,
            SheetParseValues.CALCULATIONS_START_COL + num_regions).value = label


def write_treatment_labels(outsheet: Worksheet, treatment_labels: TreatmentLabels) -> None: 
    '''
    Writes treatment_labels to the sheet

    @param outsheet: worksheet to write treatment_labels to. @param
    treatment_labels: treatment labels (CCK, etc.).
    '''
    for label in treatment_labels:
        outsheet.cell(label.index, SheetParseValues.TREATMENT_LABEL_COLUMN).value = label.label


def write_calculations_to_outsheet(
    outsheet: Worksheet,
    peak: float,
    base: float,
    std: float,
    area: float,
    delta: float,
    num_regions: int,
    curr_region_idx: int,
    treatment_name: str,
    treatment_idx) -> None:
    calculations = [treatment_name, peak, base, std, area, delta]

    for calculation_idx, calc in enumerate(calculations):
        outsheet.cell(
            SheetParseValues.CALCULATIONS_START_ROW + calculation_idx + treatment_idx * (len(calculations) + 1),
            SheetParseValues.CALCULATIONS_START_COL + num_regions + curr_region_idx + 1).value = calc


def write_data_to_outsheet(
    outsheet: Worksheet,
    concentration_data: CalciumDataRegionToTreatmentsDict) -> None:
    '''
    Populates the time column with the correct times in the new sheet. Also
    writes concentration data to the worksheet.

    @param outsheet: Worksheet to populate data with. @param time_data: List of
    time points to populate the new sheet with. @param concentration_data: Data
    structure containing all concentration data.
    '''
    TIME_COLUMN_OFFSET = 3

    outsheet.cell(1, 1, "Time (sec)")

    for i, (region_name, concentrations) in enumerate(concentration_data.items()):
        outsheet.cell(1, i + TIME_COLUMN_OFFSET).value = region_name

        j = SheetParseValues.START_DATA_ROW

        for data in concentrations:
            if type(data) is StandardBathData:
                data = data.std_bath_data
            else:
                data = data.treatment_data                    

            for value in data:
                outsheet.cell(j, 1).value = value.time
                outsheet.cell(j, i + TIME_COLUMN_OFFSET).value = value.value
                j += 1


def search_for_data_bounds(insheet: Worksheet) -> Tuple[int, int, int, int]:
    '''
    Searches for the start of the data by doing the following: - Searches for
    the "Time" column
    
    Searches for the the start of the data by looking for the first non-zero
    number in the first column but takes into account clock resets
    '''
    # Find column with title "Time" in first row
    time_column_idx = None
    start_region_idx = None
    end_region_idx = None

    for col in insheet.columns:
        if col[0].value is None:
            continue

        if time_column_idx is None and 'Time' in col[0].value:
            time_column_idx = col[0].col_idx

        if start_region_idx is None and is_valid_ratio_data_header(col[0].value):
            start_region_idx = col[0].col_idx
        
        if end_region_idx is None and \
            start_region_idx is not None and \
            not is_valid_ratio_data_header(col[0].value):
            end_region_idx = col[0].col_idx - 1

    if end_region_idx is None:
        end_region_idx = insheet.max_column

    return SheetParseValues.START_DATA_ROW, time_column_idx, start_region_idx, end_region_idx


def main():
    args = parse_arg()
    name_args = parse_data_name(args.file)    
    outfile = generate_outfilename(args.file)

    print(f'Converting calcium image file named {args.file} with the following parameters:\n\
        * output filename: {outfile}\n\
        * base: {args.base}\n\
        * peak: {args.peak}\n\
        * run: {name_args["run"]}')

    convert_calcium_file(args.file, outfile, name_args['run'], args.base, args.peak, args.post_std_time_to_search)


if __name__ == '__main__':
    main()
